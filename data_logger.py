"""
data_logger.py — Duomenų išsaugojimas į .xlsx (Excel).

Vienas failas su dviem lapais:
  Kadrai — kiekvieno kadro duomenys
  Metimai — suvestinė po kiekvieno meto
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Stulpelių apibrėžimai ──────────────────────────────────────────────────────
FRAME_COLS = [
    ('Time (s)', 'timestamp_s'),
    ('Frame No.', 'frame_idx'),
    ('Shot No.', 'shot_num'),
    ('Phase', 'phase'),
    ('L.Shoulder X', 'l_shoulder_x'),
    ('L.Shoulder Y', 'l_shoulder_y'),
    ('L.Elbow X', 'l_elbow_x'),
    ('L.Elbow Y', 'l_elbow_y'),
    ('L.Wrist X', 'l_wrist_x'),
    ('L.Wrist Y', 'l_wrist_y'),
    ('L.Hip X', 'l_hip_x'),
    ('L.Hip Y', 'l_hip_y'),
    ('L.Knee X', 'l_knee_x'),
    ('L.Knee Y', 'l_knee_y'),
    ('L.Ankle X', 'l_ankle_x'),
    ('L.Ankle Y', 'l_ankle_y'),
    ('R.Shoulder X', 'r_shoulder_x'),
    ('R.Shoulder Y', 'r_shoulder_y'),
    ('R.Elbow X', 'r_elbow_x'),
    ('R.Elbow Y', 'r_elbow_y'),
    ('R.Wrist X', 'r_wrist_x'),
    ('R.Wrist Y', 'r_wrist_y'),
    ('R.Hip X', 'r_hip_x'),
    ('R.Hip Y', 'r_hip_y'),
    ('R.Knee X', 'r_knee_x'),
    ('R.Knee Y', 'r_knee_y'),
    ('R.Ankle X', 'r_ankle_x'),
    ('R.Ankle Y', 'r_ankle_y'),
    ('Elbow Angle (°)', 'elbow_angle'),
    ('Release Angle (°)', 'release_angle'),
    ('Knee Angle (°)', 'knee_angle'),
    ('Ankle Angle (°)', 'ankle_angle'),
    ('Wrist Above Elbow (px)', 'wrist_above_elbow'),
    ('Ball State', 'ball_state'),
    ('Pose Correct', 'pose_correct'),
    ('Shot Result', 'shot_result'),
]

SHOT_COLS = [
    ('Shot No.', 'shot_num'),
    ('Result', 'result'),
    ('Start (s)', 'start_time_s'),
    ('End (s)', 'end_time_s'),
    ('Duration (s)', 'duration_s'),
    ('Max Elbow Angle (°)', 'max_elbow_angle'),
    ('Max Release Angle (°)', 'max_release_angle'),
    ('Min Knee Angle (°)', 'min_knee_angle'),
    ('Elbow Flare?', 'elbow_flare_detected'),
    ('Follow-through?', 'no_followthrough_detected'),
    ('Knee Issue?', 'knee_issue_detected'),
    ('Low Release?', 'release_low_detected'),
    ('Frame Count', 'frame_count'),
]


# ── Stiliai ───────────────────────────────────────────────────────────────────
HDR_FILL_FRAMES = PatternFill('solid', start_color = '1F4E79')
HDR_FILL_SHOTS  = PatternFill('solid', start_color = '375623')
HDR_FONT = Font(bold = True, color = 'FFFFFF', name = 'Arial', size = 10)
DATA_FONT = Font(name = 'Arial', size = 9)
GOOD_FILL = PatternFill('solid', start_color = 'C6EFCE')
BAD_FILL = PatternFill('solid', start_color = 'FFC7CE')
CENTER = Alignment(horizontal = 'center', vertical = 'center')
LEFT = Alignment(horizontal = 'left', vertical = 'center')
THIN_BORDER = Border(
    left = Side(style = 'thin', color = 'CCCCCC'),
    right = Side(style = 'thin', color = 'CCCCCC'),
    top = Side(style = 'thin', color = 'CCCCCC'),
    bottom = Side(style = 'thin', color = 'CCCCCC'),
)


class DataLogger:
    def __init__(self, output_dir='data'):
        os.makedirs(output_dir, exist_ok=True)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')

        self._path = os.path.join(output_dir, f'basketball_{ts}.xlsx')
        self._wb = Workbook()
        self._ws_frames = self._wb.active
        self._ws_frames.title = 'Frames'
        self._ws_shots = self._wb.create_sheet('Shots')
        self._setup_headers(self._ws_frames, FRAME_COLS, HDR_FILL_FRAMES)
        self._setup_headers(self._ws_shots,  SHOT_COLS,  HDR_FILL_SHOTS)

        self._frame_idx = 0
        self._shot_buf = self._new_shot_buf()
        self._frames_row = 2
        self._shots_row = 2

        self._flush_every = 30
        self._flush_cnt = 0

        print(f'[LOG] File: {self._path}')

    # ── Antraštės ─────────────────────────────────────────────────────────────
    @staticmethod
    def _setup_headers(ws, col_defs, fill):
        for col_idx, (label, _) in enumerate(col_defs, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font      = HDR_FONT
            cell.fill      = fill
            cell.alignment = CENTER
            cell.border    = THIN_BORDER
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = 'A2'

    @staticmethod
    def _auto_width(ws, col_defs):
        for col_idx, (label, _) in enumerate(col_defs, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(len(label) + 2, 10)

    # ── Buferis ───────────────────────────────────────────────────────────────
    @staticmethod
    def _new_shot_buf():
        return {
            'max_elbow_angle'  : 0,
            'max_release_angle': 0,
            'min_knee_angle'   : 999,
            'elbow_flare'      : False,
            'no_followthrough' : False,
            'knee_issue'       : False,
            'release_low'      : False,
            'frame_count'      : 0,
            'start_time_s'     : None,
        }

    # ── Kadrų išsaugojimas ────────────────────────────────────────────────────
    def log_frame(self, timestamp_s, shot_num, phase, landmarks, angles,
                  ball_state, pose_correct, feedback_flags, shot_result=None):
        lm = landmarks
        a = angles

        values = {
            'timestamp_s': round(timestamp_s, 3),
            'frame_idx': self._frame_idx,
            'shot_num': shot_num,
            'phase': phase or 'UNKNOWN',
            'l_shoulder_x': lm.get('l_shoulder', ['',''])[0],
            'l_shoulder_y': lm.get('l_shoulder', ['',''])[1],
            'l_elbow_x': lm.get('l_elbow', ['',''])[0],
            'l_elbow_y': lm.get('l_elbow', ['',''])[1],
            'l_wrist_x': lm.get('l_wrist', ['',''])[0],
            'l_wrist_y': lm.get('l_wrist', ['',''])[1],
            'l_hip_x': lm.get('l_hip', ['',''])[0],
            'l_hip_y': lm.get('l_hip', ['',''])[1],
            'l_knee_x': lm.get('l_knee', ['',''])[0],
            'l_knee_y': lm.get('l_knee', ['',''])[1],
            'l_ankle_x': lm.get('l_ankle', ['',''])[0],
            'l_ankle_y': lm.get('l_ankle', ['',''])[1],
            'r_shoulder_x': lm.get('r_shoulder', ['',''])[0],
            'r_shoulder_y': lm.get('r_shoulder', ['',''])[1],
            'r_elbow_x': lm.get('r_elbow', ['',''])[0],
            'r_elbow_y': lm.get('r_elbow', ['',''])[1],
            'r_wrist_x': lm.get('r_wrist', ['',''])[0],
            'r_wrist_y': lm.get('r_wrist', ['',''])[1],
            'r_hip_x': lm.get('r_hip', ['',''])[0],
            'r_hip_y': lm.get('r_hip', ['',''])[1],
            'r_knee_x': lm.get('r_knee', ['',''])[0],
            'r_knee_y': lm.get('r_knee', ['',''])[1],
            'r_ankle_x': lm.get('r_ankle', ['',''])[0],
            'r_ankle_y': lm.get('r_ankle', ['',''])[1],
            'elbow_angle': a.get('elbow', ''),
            'release_angle': a.get('release', ''),
            'knee_angle': a.get('knee', ''),
            'ankle_angle' : a.get('ankle', ''),
            'wrist_above_elbow': a.get('wrist_above', ''),
            'ball_state': ball_state,
            'pose_correct': 'YES' if pose_correct else 'NO',
            'shot_result': shot_result if shot_result else '-',
        }

        row = self._frames_row
        for col_idx, (_, key) in enumerate(FRAME_COLS, start = 1):
            cell = self._ws_frames.cell(row = row, column = col_idx, value = values.get(key, ''))
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER if col_idx > 4 else LEFT
            if key == 'pose_correct':
                cell.fill = GOOD_FILL if pose_correct else BAD_FILL
            if key == 'shot_result':
                if values.get('shot_result') == 'GOOD':
                    cell.fill = GOOD_FILL
                elif values.get('shot_result') == 'BAD':
                    cell.fill = BAD_FILL

        self._frames_row += 1
        self._frame_idx += 1

        # Periodiškai išsaugome
        self._flush_cnt += 1
        if self._flush_cnt >= self._flush_every:
            self._wb.save(self._path)
            self._flush_cnt = 0

        # Kaupiame shot buffer
        buf = self._shot_buf
        buf['frame_count'] += 1
        if buf['start_time_s'] is None:
            buf['start_time_s'] = timestamp_s
        buf['max_elbow_angle'] = max(buf['max_elbow_angle'], a.get('elbow', 0) or 0)
        buf['max_release_angle'] = max(buf['max_release_angle'], a.get('release', 0) or 0)
        knee = a.get('knee', 999) or 999
        if knee < buf['min_knee_angle']:
            buf['min_knee_angle'] = knee
        if len(feedback_flags) >= 4:
            if feedback_flags[0]: buf['elbow_flare'] = True
            if feedback_flags[1]: buf['no_followthrough'] = True
            if feedback_flags[2]: buf['knee_issue'] = True
            if feedback_flags[3]: buf['release_low'] = True

    # ── Metų išsaugojimas ─────────────────────────────────────────────────────
    def log_shot(self, shot_num, result, end_time_s, shot_errors=None):
        buf   = self._shot_buf
        start = buf['start_time_s'] if buf['start_time_s'] is not None else end_time_s
        errs  = shot_errors or set()

        values = {
            'shot_num': shot_num,
            'result': result,
            'start_time_s': round(start, 3),
            'end_time_s': round(end_time_s, 3),
            'duration_s': round(end_time_s - start, 3),
            'max_elbow_angle': buf['max_elbow_angle'],
            'max_release_angle' : buf['max_release_angle'],
            'min_knee_angle': buf['min_knee_angle'] if buf['min_knee_angle'] < 999 else '',
            'elbow_flare_detected': 'YES' if 0 in errs else 'NO',
            'no_followthrough_detected': 'YES' if 1 in errs else 'NO',
            'knee_issue_detected': 'YES' if 2 in errs else 'NO',
            'release_low_detected': 'YES' if 3 in errs else 'NO',
            'frame_count': buf['frame_count'],
        }

        row = self._shots_row
        is_good = (result == 'GOOD')
        for col_idx, (_, key) in enumerate(SHOT_COLS, start=1):
            cell = self._ws_shots.cell(row = row, column = col_idx, value = values.get(key, ''))
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            if key == 'result':
                cell.fill = GOOD_FILL if is_good else BAD_FILL
            if key in ('elbow_flare_detected', 'no_followthrough_detected',
                       'knee_issue_detected', 'release_low_detected'):
                if values.get(key) == 'TAIP':
                    cell.fill = BAD_FILL

        self._shots_row += 1
        self._shot_buf = self._new_shot_buf()
        self._wb.save(self._path)
        print(f'[LOG] Shot #{shot_num} → {result} | duration {values["duration_s"]}s')

    # ── Tuščias metodai suderinamumui su main.py ──────────────────────────────
    def start_shot(self):
        pass

    def end_shot(self):
        pass

    def reset_shot_buf(self):
        self._shot_buf = self._new_shot_buf()

    # ── Uždarymas ─────────────────────────────────────────────────────────────
    def close(self):
        self._auto_width(self._ws_frames, FRAME_COLS)
        self._auto_width(self._ws_shots,  SHOT_COLS)
        if self._shots_row > 2:
            last = self._shots_row - 1
            sum_row = self._shots_row + 1
            self._ws_shots.cell(row = sum_row, column = 1, value='TOTAL').font = Font(bold = True, name = 'Arial')
            self._ws_shots.cell(row = sum_row, column = 2,
                value=f'=COUNTIF(B2:B{last},"GOOD")&" GOOD / "&COUNTIF(B2:B{last},"BAD")&" BAD"')
        self._wb.save(self._path)
        print(f'[LOG] File saved: {self._path}')